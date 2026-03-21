from __future__ import annotations

import base64
import csv
import json
import mimetypes
import os
from datetime import date, datetime, time
from decimal import Decimal
from io import StringIO
from typing import Any, Dict, List, Optional, Sequence

from ..agent_types import AgentTool, AgentToolResult, TextContent
from .path_roots import normalize_allowed_roots, path_in_roots

from pypdf import PdfReader
from openpyxl import load_workbook



_TEXT_SUFFIXES = {
    ".txt",
    ".md",
    ".log",
    ".rst",
    ".jsonl",
    ".yaml",
    ".yml",
    ".toml",
    ".ini",
    ".cfg",
    ".conf",
    ".py",
    ".pyi",
    ".sh",
    ".bash",
    ".zsh",
    ".ps1",
    ".bat",
    ".cmd",
    ".js",
    ".jsx",
    ".ts",
    ".tsx",
    ".css",
    ".scss",
    ".html",
    ".htm",
    ".xml",
    ".sql",
    ".tex",
    ".pdb",
    ".cif",
    ".mmcif",
}
_STRUCTURE_TEXT_SUFFIXES = {
    ".pdb",
    ".cif",
    ".mmcif",
}
_IMAGE_SUFFIXES = {
    ".png",
    ".jpg",
    ".jpeg",
    ".gif",
    ".bmp",
    ".webp",
    ".tif",
    ".tiff",
    ".svg",
    ".ico",
    ".heic",
    ".heif",
    ".avif",
}


class FileReadError(RuntimeError):
    def __init__(self, code: str, message: str) -> None:
        super().__init__(str(message))
        self.code = str(code or "parse_error")
        self.message = str(message or "unknown file read error")


def _success_payload(
    *,
    path: str,
    workspace_path: str,
    suffix: str,
    kind: str,
    content: Any,
    metadata: Dict[str, Any],
    truncated: bool,
) -> Dict[str, Any]:
    return {
        "ok": True,
        "path": path,
        "workspace_path": workspace_path,
        "suffix": suffix,
        "kind": kind,
        "metadata": metadata,
        "truncated": truncated,
        "content": content,
    }


def _resolve_workspace_path(path_value: str, workspace_root: str, allowed_roots: Sequence[str]) -> str:
    candidate = str(path_value or "").strip()
    if not candidate:
        raise FileReadError("path_not_found", "path is required")
    absolute = os.path.realpath(candidate if os.path.isabs(candidate) else os.path.join(workspace_root, candidate))
    if not path_in_roots(absolute, allowed_roots):
        raise FileReadError("path_outside_workspace", "path must be inside workspace or configured allowed roots")
    if not os.path.exists(absolute):
        raise FileReadError("path_not_found", f"file not found: {path_value}")
    if not os.path.isfile(absolute):
        raise FileReadError("path_not_file", f"path is not a file: {path_value}")
    return absolute


def _workspace_relative_path(path: str, workspace_root: str) -> str:
    try:
        rel = os.path.relpath(path, workspace_root)
    except Exception:
        rel = path
    return str(rel).replace("\\", "/")


def _read_utf8_text(path: str) -> str:
    try:
        with open(path, "r", encoding="utf-8") as handle:
            return handle.read()
    except UnicodeDecodeError as exc:
        raise FileReadError("parse_error", f"unable to decode UTF-8 text: {exc}") from exc
    except OSError as exc:
        raise FileReadError("parse_error", f"unable to read file: {exc}") from exc


def _detect_image_kind(path: str, suffix: str) -> bool:
    if suffix in _IMAGE_SUFFIXES:
        return True
    mime_type, _ = mimetypes.guess_type(path)
    return str(mime_type or "").strip().lower().startswith("image/")


def _coerce_optional_positive_int(value: Any, field_name: str) -> Optional[int]:
    if value in (None, ""):
        return None
    try:
        parsed = int(value)
    except Exception as exc:
        raise FileReadError("parse_error", f"{field_name} must be an integer") from exc
    if parsed <= 0:
        raise FileReadError("parse_error", f"{field_name} must be > 0")
    return parsed


def _slice_text_lines(
    lines: Sequence[str],
    *,
    line_start: Optional[int],
    line_count: Optional[int],
) -> tuple[str, Dict[str, Any]]:
    total_lines = len(lines)
    start = int(line_start or 1)
    start_index = min(max(start - 1, 0), total_lines)
    if line_count is None:
        end_index = total_lines
    else:
        end_index = min(total_lines, start_index + int(line_count))

    selected = "\n".join(lines[start_index:end_index])
    has_selection = start_index < end_index
    return selected, {
        "line_count": total_lines,
        "requested_line_start": start,
        "requested_line_count": int(line_count) if line_count is not None else None,
        "returned_line_start": start_index + 1 if has_selection else 0,
        "returned_line_end": end_index if has_selection else 0,
    }


def _build_structure_text_metadata(lines: Sequence[str], suffix: str) -> Dict[str, Any]:
    metadata: Dict[str, Any] = {"line_count": len(lines)}
    if suffix == ".pdb":
        atom_preview: List[str] = []
        first_atom_line: Optional[int] = None
        first_hetatm_line: Optional[int] = None
        first_model_line: Optional[int] = None
        for index, line in enumerate(lines, start=1):
            if first_model_line is None and line.startswith("MODEL"):
                first_model_line = index
            if line.startswith("ATOM"):
                if first_atom_line is None:
                    first_atom_line = index
                if len(atom_preview) < 3:
                    atom_preview.append(line.rstrip())
            elif first_hetatm_line is None and line.startswith("HETATM"):
                first_hetatm_line = index
            if first_atom_line is not None and len(atom_preview) >= 3 and first_hetatm_line is not None:
                break
        if first_atom_line is not None:
            metadata["first_atom_line"] = first_atom_line
        if atom_preview:
            metadata["atom_preview"] = atom_preview
        if first_hetatm_line is not None:
            metadata["first_hetatm_line"] = first_hetatm_line
        if first_model_line is not None:
            metadata["first_model_line"] = first_model_line
        return metadata

    atom_site_preview: List[str] = []
    first_atom_site_line: Optional[int] = None
    first_loop_line: Optional[int] = None
    for index, line in enumerate(lines, start=1):
        stripped = line.strip()
        if first_loop_line is None and stripped == "loop_":
            first_loop_line = index
        if stripped.startswith("_atom_site."):
            if first_atom_site_line is None:
                first_atom_site_line = index
            if len(atom_site_preview) < 3:
                atom_site_preview.append(stripped)
        if first_atom_site_line is not None and len(atom_site_preview) >= 3 and first_loop_line is not None:
            break
    if first_loop_line is not None:
        metadata["first_loop_line"] = first_loop_line
    if first_atom_site_line is not None:
        metadata["first_atom_site_line"] = first_atom_site_line
    if atom_site_preview:
        metadata["atom_site_preview"] = atom_site_preview
    return metadata


def _safe_xlsx_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (date, datetime, time)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, bytes):
        return base64.b64encode(value).decode("ascii")
    return str(value)


def _sanitize_headers(header_row: Any) -> List[str]:
    names: List[str] = []
    counts: Dict[str, int] = {}
    for index, raw in enumerate(tuple(header_row or ()), start=1):
        base = str(raw).strip() if raw is not None else ""
        if not base:
            base = f"column_{index}"
        seen = counts.get(base, 0)
        counts[base] = seen + 1
        names.append(base if seen == 0 else f"{base}_{seen + 1}")
    return names


def _xlsx_sheet_to_rows(sheet: Any) -> tuple[List[str], List[Dict[str, Any]]]:
    values = [tuple(row) for row in sheet.iter_rows(values_only=True)]
    if not values:
        return [], []

    headers = _sanitize_headers(values[0])
    rows: List[Dict[str, Any]] = []
    for row in values[1:]:
        width = max(len(headers), len(row))
        if width == 0:
            continue
        if len(headers) < width:
            start = len(headers)
            for offset in range(start, width):
                headers.append(f"column_{offset + 1}")
        normalized = [_safe_xlsx_value(row[idx]) if idx < len(row) else None for idx in range(width)]
        if not any(cell not in (None, "") for cell in normalized):
            continue
        record = {headers[idx]: normalized[idx] for idx in range(width)}
        rows.append(record)
    return headers, rows


def _xlsx_to_json_payload(path: str, *, byte_size: int) -> Dict[str, Any]:
    if load_workbook is None:
        raise FileReadError("missing_dependency", "openpyxl is required to read xlsx files")

    try:
        workbook = load_workbook(path, data_only=True, read_only=True)
    except FileReadError:
        raise
    except Exception as exc:
        raise FileReadError("parse_error", f"unable to parse xlsx: {exc}") from exc

    try:
        sheets_payload: List[Dict[str, Any]] = []
        sheets_meta: List[Dict[str, Any]] = []
        for sheet in workbook.worksheets:
            headers, rows = _xlsx_sheet_to_rows(sheet)
            sheets_payload.append({"name": sheet.title, "headers": headers, "rows": rows})
            sheets_meta.append({"name": sheet.title, "headers": headers, "row_count": len(rows)})
    finally:
        try:
            workbook.close()
        except Exception:
            pass

    total_rows = sum(item["row_count"] for item in sheets_meta)
    return {
        "content": {"sheets": sheets_payload},
        "metadata": {
            "format": "json",
            "sheet_count": len(sheets_meta),
            "total_row_count": total_rows,
            "sheets": sheets_meta,
            "byte_size": byte_size,
        },
    }


def _xlsx_to_csv_payload(path: str, *, byte_size: int) -> Dict[str, Any]:
    if load_workbook is None:
        raise FileReadError("missing_dependency", "openpyxl is required to read xlsx files")

    try:
        workbook = load_workbook(path, data_only=True, read_only=True)
    except FileReadError:
        raise
    except Exception as exc:
        raise FileReadError("parse_error", f"unable to parse xlsx: {exc}") from exc

    try:
        parts: List[str] = []
        sheets_meta: List[Dict[str, Any]] = []
        include_sheet_banner = len(workbook.worksheets) > 1
        for sheet in workbook.worksheets:
            headers, rows = _xlsx_sheet_to_rows(sheet)
            buffer = StringIO()
            if headers:
                writer = csv.DictWriter(buffer, fieldnames=headers)
                writer.writeheader()
                writer.writerows(rows)
            sheet_csv = buffer.getvalue().rstrip("\n")
            if include_sheet_banner:
                parts.append(f"# sheet: {sheet.title}")
            parts.append(sheet_csv)
            sheets_meta.append({"name": sheet.title, "headers": headers, "row_count": len(rows)})
    finally:
        try:
            workbook.close()
        except Exception:
            pass

    text = "\n\n".join(parts).strip()
    return {
        "content": text,
        "metadata": {
            "format": "csv",
            "sheet_count": len(sheets_meta),
            "total_row_count": sum(item["row_count"] for item in sheets_meta),
            "sheets": sheets_meta,
            "char_count": len(text),
            "byte_size": byte_size,
        },
    }


def read(
    path: str,
    *,
    workspace_root: Optional[str] = None,
    allowed_roots: Optional[Sequence[str]] = None,
    max_chars: int = 200_000,
    max_bytes: int = 10_000_000,
    xlsx_format: str = "json",
    line_start: Optional[int] = None,
    line_count: Optional[int] = None,
) -> Dict[str, Any]:
    root, roots = normalize_allowed_roots(workspace_root, allowed_roots)
    try:
        max_chars_value = int(max_chars)
    except Exception as exc:
        raise FileReadError("parse_error", "max_chars must be an integer") from exc
    if max_chars_value <= 0:
        raise FileReadError("parse_error", "max_chars must be > 0")

    try:
        max_bytes_value = int(max_bytes)
    except Exception as exc:
        raise FileReadError("parse_error", "max_bytes must be an integer") from exc
    if max_bytes_value <= 0:
        raise FileReadError("parse_error", "max_bytes must be > 0")

    line_start_value = _coerce_optional_positive_int(line_start, "line_start")
    line_count_value = _coerce_optional_positive_int(line_count, "line_count")

    absolute = _resolve_workspace_path(path, root, roots)
    size = os.path.getsize(absolute)
    if size > max_bytes_value:
        raise FileReadError("file_too_large", f"file size {size} exceeds max_bytes={max_bytes_value}")

    suffix = os.path.splitext(absolute)[1].lower()
    workspace_path = _workspace_relative_path(absolute, root)

    if suffix in _TEXT_SUFFIXES:
        text = _read_utf8_text(absolute)
        original_chars = len(text)
        metadata: Dict[str, Any] = {"encoding": "utf-8", "char_count": original_chars, "byte_size": size}
        selected_text = text
        if line_start_value is not None or line_count_value is not None or suffix in _STRUCTURE_TEXT_SUFFIXES:
            lines = text.splitlines()
            if suffix in _STRUCTURE_TEXT_SUFFIXES:
                metadata.update(_build_structure_text_metadata(lines, suffix))
            elif "line_count" not in metadata:
                metadata["line_count"] = len(lines)
            if line_start_value is not None or line_count_value is not None:
                selected_text, line_metadata = _slice_text_lines(
                    lines,
                    line_start=line_start_value,
                    line_count=line_count_value,
                )
                metadata.update(line_metadata)
        truncated = len(selected_text) > max_chars_value
        return _success_payload(
            path=absolute,
            workspace_path=workspace_path,
            suffix=suffix,
            kind="text",
            content=selected_text[:max_chars_value] if truncated else selected_text,
            metadata=metadata,
            truncated=truncated,
        )

    if suffix == ".json":
        text = _read_utf8_text(absolute)
        try:
            parsed = json.loads(text)
        except Exception as exc:
            raise FileReadError("parse_error", f"invalid json: {exc}") from exc
        return _success_payload(
            path=absolute,
            workspace_path=workspace_path,
            suffix=suffix,
            kind="json",
            content=parsed,
            metadata={"encoding": "utf-8", "byte_size": size, "value_type": type(parsed).__name__},
            truncated=False,
        )

    if suffix == ".csv":
        text = _read_utf8_text(absolute)
        try:
            reader = csv.DictReader(StringIO(text))
            headers = list(reader.fieldnames or [])
            rows = [dict(row) for row in reader]
        except Exception as exc:
            raise FileReadError("parse_error", f"invalid csv: {exc}") from exc
        return _success_payload(
            path=absolute,
            workspace_path=workspace_path,
            suffix=suffix,
            kind="csv",
            content=rows,
            metadata={"encoding": "utf-8", "headers": headers, "row_count": len(rows), "byte_size": size},
            truncated=False,
        )

    if suffix == ".pdf":
        if PdfReader is None:
            raise FileReadError("missing_dependency", "pypdf is required to read pdf files")
        try:
            with open(absolute, "rb") as handle:
                reader = PdfReader(handle)
                pages = list(reader.pages)
                parts = []
                non_empty_pages = 0
                for page in pages:
                    extracted = str(page.extract_text() or "")
                    if extracted.strip():
                        non_empty_pages += 1
                    parts.append(extracted)
        except FileReadError:
            raise
        except Exception as exc:
            raise FileReadError("parse_error", f"unable to parse pdf: {exc}") from exc

        text = "\n\n".join(parts)
        original_chars = len(text)
        truncated = original_chars > max_chars_value
        return _success_payload(
            path=absolute,
            workspace_path=workspace_path,
            suffix=suffix,
            kind="pdf",
            content=text[:max_chars_value] if truncated else text,
            metadata={
                "page_count": len(pages),
                "non_empty_pages": non_empty_pages,
                "extracted_char_count": original_chars,
                "byte_size": size,
            },
            truncated=truncated,
        )

    if suffix == ".xlsx":
        normalized_xlsx_format = str(xlsx_format or "json").strip().lower()
        if normalized_xlsx_format not in {"json", "csv"}:
            raise FileReadError("parse_error", "xlsx_format must be one of: json, csv")
        converter = _xlsx_to_json_payload if normalized_xlsx_format == "json" else _xlsx_to_csv_payload
        payload = converter(absolute, byte_size=size)
        return _success_payload(
            path=absolute,
            workspace_path=workspace_path,
            suffix=suffix,
            kind="xlsx",
            content=payload["content"],
            metadata=payload["metadata"],
            truncated=False,
        )

    if _detect_image_kind(absolute, suffix):
        try:
            with open(absolute, "rb") as handle:
                payload = handle.read()
        except OSError as exc:
            raise FileReadError("parse_error", f"unable to read image: {exc}") from exc
        mime_type, _ = mimetypes.guess_type(absolute)
        encoded = base64.b64encode(payload).decode("ascii")
        return _success_payload(
            path=absolute,
            workspace_path=workspace_path,
            suffix=suffix,
            kind="image",
            content=encoded,
            metadata={
                "mime_type": str(mime_type or "application/octet-stream"),
                "byte_size": len(payload),
                "base64_chars": len(encoded),
            },
            truncated=False,
        )

    raise FileReadError("unsupported_suffix", f"unsupported file suffix: {suffix or '(none)'}")


class FileReadTool(AgentTool):
    name = "file_read"
    description = (
        "Read and parse files by suffix inside the workspace or allowed roots. "
        "Supports common text/code/config files, structure text files like pdb/cif, plus json, csv, xlsx, pdf, and image-to-base64. "
        "For large text files, can return bounded line windows and structure hints."
    )
    parameters = {
        "type": "object",
        "properties": {
            "path": {
                "type": "string",
                "description": "File path relative to the workspace root or an absolute path inside allowed roots.",
            },
            "max_chars": {"type": "integer", "description": "Optional max returned chars for text/pdf content."},
            "line_start": {
                "type": "integer",
                "description": "Optional 1-based start line for supported text-like files.",
            },
            "line_count": {
                "type": "integer",
                "description": "Optional number of lines to return for supported text-like files.",
            },
            "xlsx_format": {
                "type": "string",
                "enum": ["json", "csv"],
                "description": "Optional .xlsx conversion format. Defaults to json.",
            },
        },
        "required": ["path"],
    }
    label = "File Read"

    def __init__(
        self,
        workspace_root: Optional[str] = None,
        max_output_chars: int = 20_000,
        allowed_roots: Optional[Sequence[str]] = None,
    ) -> None:
        self.workspace_root, self.allowed_roots = normalize_allowed_roots(workspace_root, allowed_roots)
        self.max_output_chars = max(512, int(max_output_chars or 20_000))

    @staticmethod
    def _error(path_value: str, code: str, message: str) -> AgentToolResult:
        return AgentToolResult(
            content=[TextContent(type="text", text=f"file_read error: {message}")],
            details={
                "ok": False,
                "error": str(code or "parse_error"),
                "message": str(message or "file read failed"),
                "path": str(path_value or ""),
            },
        )

    async def execute(
        self,
        tool_call_id: str,
        params: Dict[str, Any],
        signal: Optional[Any] = None,
        on_update: Optional[Any] = None,
    ) -> AgentToolResult:
        del tool_call_id, signal, on_update
        path_value = str(params.get("path", "") or "").strip()
        if not path_value:
            return self._error(path_value, "path_not_found", "path is required")

        max_chars = params.get("max_chars", 200_000)
        line_start = params.get("line_start")
        line_count = params.get("line_count")
        xlsx_format = str(params.get("xlsx_format", "json") or "json")
        try:
            payload = read(
                path_value,
                workspace_root=self.workspace_root,
                allowed_roots=self.allowed_roots,
                max_chars=int(max_chars),
                xlsx_format=xlsx_format,
                line_start=line_start,
                line_count=line_count,
            )
        except FileReadError as exc:
            return self._error(path_value, exc.code, exc.message)
        except Exception as exc:
            return self._error(path_value, "parse_error", str(exc))

        details: Dict[str, Any] = dict(payload)
        rendered = json.dumps(payload, ensure_ascii=False)
        if len(rendered) > self.max_output_chars:
            rendered = rendered[: self.max_output_chars]
            details["output_truncated"] = True
            details["max_output_chars"] = self.max_output_chars

        return AgentToolResult(
            content=[TextContent(type="text", text=rendered)],
            details=details,
        )
