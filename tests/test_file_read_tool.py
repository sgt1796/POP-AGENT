from __future__ import annotations

import asyncio
import base64
import json
from pathlib import Path
import openpyxl
import pytest

from agent.tools.file_read_tool import FileReadError, FileReadTool, read


def _run(tool, params):
    return asyncio.run(tool.execute("tc1", params))


def test_read_txt_and_md(tmp_path: Path):
    (tmp_path / "notes.txt").write_text("hello text", encoding="utf-8")
    (tmp_path / "guide.md").write_text("# header", encoding="utf-8")

    txt = read("notes.txt", workspace_root=str(tmp_path))
    md = read("guide.md", workspace_root=str(tmp_path))

    assert txt["ok"] is True
    assert txt["kind"] == "text"
    assert txt["content"] == "hello text"
    assert txt["suffix"] == ".txt"

    assert md["ok"] is True
    assert md["kind"] == "text"
    assert md["content"] == "# header"
    assert md["suffix"] == ".md"


def test_read_common_text_like_suffixes(tmp_path: Path):
    (tmp_path / "notes.rst").write_text("section", encoding="utf-8")
    (tmp_path / "config.yaml").write_text("mode: fast", encoding="utf-8")
    (tmp_path / "events.jsonl").write_text('{"step": 1}\n{"step": 2}\n', encoding="utf-8")
    (tmp_path / "structure.pdb").write_text("ATOM      1  N   GLY A   1      11.104  13.207   9.199\n", encoding="utf-8")
    (tmp_path / "structure.cif").write_text("data_demo\n_cell.length_a 10.0\n", encoding="utf-8")
    (tmp_path / "structure.mmcif").write_text("data_demo\n_cell.length_b 11.0\n", encoding="utf-8")

    rst = read("notes.rst", workspace_root=str(tmp_path))
    yaml = read("config.yaml", workspace_root=str(tmp_path))
    jsonl = read("events.jsonl", workspace_root=str(tmp_path))
    pdb = read("structure.pdb", workspace_root=str(tmp_path))
    cif = read("structure.cif", workspace_root=str(tmp_path))
    mmcif = read("structure.mmcif", workspace_root=str(tmp_path))

    assert rst["ok"] is True
    assert rst["kind"] == "text"
    assert rst["content"] == "section"
    assert rst["suffix"] == ".rst"

    assert yaml["ok"] is True
    assert yaml["kind"] == "text"
    assert yaml["content"] == "mode: fast"
    assert yaml["suffix"] == ".yaml"

    assert jsonl["ok"] is True
    assert jsonl["kind"] == "text"
    assert '{"step": 1}' in jsonl["content"]
    assert jsonl["suffix"] == ".jsonl"

    assert pdb["ok"] is True
    assert pdb["kind"] == "text"
    assert "ATOM" in pdb["content"]
    assert pdb["suffix"] == ".pdb"

    assert cif["ok"] is True
    assert cif["kind"] == "text"
    assert "_cell.length_a" in cif["content"]
    assert cif["suffix"] == ".cif"

    assert mmcif["ok"] is True
    assert mmcif["kind"] == "text"
    assert "_cell.length_b" in mmcif["content"]
    assert mmcif["suffix"] == ".mmcif"


def test_read_json(tmp_path: Path):
    (tmp_path / "payload.json").write_text('{"a": 1, "b": "two"}', encoding="utf-8")

    result = read("payload.json", workspace_root=str(tmp_path))

    assert result["ok"] is True
    assert result["kind"] == "json"
    assert result["content"] == {"a": 1, "b": "two"}
    assert result["truncated"] is False


def test_read_csv(tmp_path: Path):
    (tmp_path / "table.csv").write_text("name,age\nAlice,30\nBob,31\n", encoding="utf-8")

    result = read("table.csv", workspace_root=str(tmp_path))

    assert result["ok"] is True
    assert result["kind"] == "csv"
    assert result["metadata"]["headers"] == ["name", "age"]
    assert result["metadata"]["row_count"] == 2
    assert result["content"] == [
        {"name": "Alice", "age": "30"},
        {"name": "Bob", "age": "31"},
    ]


def test_read_xlsx_json(tmp_path: Path):
    workbook = openpyxl.Workbook()
    movies = workbook.active
    movies.title = "Movies"
    movies.append(["title", "year", "in_stock"])
    movies.append(["Inception", 2010, True])
    movies.append(["Spirited Away", 2001, False])
    directors = workbook.create_sheet("Directors")
    directors.append(["name"])
    directors.append(["Christopher Nolan"])
    workbook.save(tmp_path / "inventory.xlsx")

    result = read("inventory.xlsx", workspace_root=str(tmp_path))

    assert result["ok"] is True
    assert result["kind"] == "xlsx"
    assert result["metadata"]["format"] == "json"
    assert result["metadata"]["sheet_count"] == 2
    assert result["metadata"]["total_row_count"] == 3
    sheets = result["content"]["sheets"]
    assert sheets[0]["name"] == "Movies"
    assert sheets[0]["rows"][0] == {"title": "Inception", "year": 2010, "in_stock": True}
    assert sheets[0]["rows"][1] == {"title": "Spirited Away", "year": 2001, "in_stock": False}
    assert sheets[1]["name"] == "Directors"
    assert sheets[1]["rows"] == [{"name": "Christopher Nolan"}]


def test_read_xlsx_csv(tmp_path: Path):
    openpyxl = pytest.importorskip("openpyxl")
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Movies"
    sheet.append(["title", "year"])
    sheet.append(["Arrival", 2016])
    workbook.save(tmp_path / "movies.xlsx")

    result = read("movies.xlsx", workspace_root=str(tmp_path), xlsx_format="csv")

    assert result["ok"] is True
    assert result["kind"] == "xlsx"
    assert result["metadata"]["format"] == "csv"
    assert result["metadata"]["sheet_count"] == 1
    assert result["metadata"]["total_row_count"] == 1
    assert "title,year" in result["content"]
    assert "Arrival,2016" in result["content"]


def test_read_xlsx_rejects_invalid_format(tmp_path: Path):
    (tmp_path / "empty.xlsx").write_bytes(b"placeholder")

    with pytest.raises(FileReadError) as exc:
        read("empty.xlsx", workspace_root=str(tmp_path), xlsx_format="xml")

    assert exc.value.code == "parse_error"


def test_read_pdf_includes_page_metadata(tmp_path: Path):
    pypdf = pytest.importorskip("pypdf")
    writer = pypdf.PdfWriter()
    writer.add_blank_page(width=72, height=72)
    pdf_path = tmp_path / "sample.pdf"
    with pdf_path.open("wb") as handle:
        writer.write(handle)

    result = read("sample.pdf", workspace_root=str(tmp_path))

    assert result["ok"] is True
    assert result["kind"] == "pdf"
    assert result["metadata"]["page_count"] == 1
    assert "extracted_char_count" in result["metadata"]
    assert isinstance(result["content"], str)


def test_read_image_as_base64(tmp_path: Path):
    png_bytes = base64.b64decode("iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAwMCAO6pYQAAAABJRU5ErkJggg==")
    image_path = tmp_path / "pixel.png"
    image_path.write_bytes(png_bytes)

    result = read("pixel.png", workspace_root=str(tmp_path))

    assert result["ok"] is True
    assert result["kind"] == "image"
    assert result["metadata"]["mime_type"] == "image/png"
    assert base64.b64decode(result["content"]) == png_bytes


def test_read_rejects_path_outside_workspace(tmp_path: Path):
    outside = tmp_path.parent / "outside_file.txt"
    outside.write_text("x", encoding="utf-8")

    with pytest.raises(FileReadError) as exc:
        read(str(outside), workspace_root=str(tmp_path))

    assert exc.value.code == "path_outside_workspace"


def test_read_allows_symlink_target_in_allowed_roots(tmp_path: Path):
    workspace = tmp_path / "workspace"
    workspace.mkdir()
    external = tmp_path / "external"
    external.mkdir()
    target = external / "nohup.log"
    target.write_text("daemon output", encoding="utf-8")
    try:
        (workspace / "Trader").symlink_to(external, target_is_directory=True)
    except OSError:
        pytest.skip("symlink creation is not supported in this environment")

    result = read(
        "Trader/nohup.log",
        workspace_root=str(workspace),
        allowed_roots=[str(external)],
    )

    assert result["ok"] is True
    assert result["content"] == "daemon output"
    assert result["path"] == str(target.resolve())


def test_read_rejects_unsupported_suffix(tmp_path: Path):
    (tmp_path / "blob.bin").write_bytes(b"\x00\x01\x02")

    with pytest.raises(FileReadError) as exc:
        read("blob.bin", workspace_root=str(tmp_path))

    assert exc.value.code == "unsupported_suffix"


def test_file_read_tool_success_envelope(tmp_path: Path):
    (tmp_path / "message.txt").write_text("hello tool", encoding="utf-8")
    tool = FileReadTool(workspace_root=str(tmp_path), max_output_chars=50_000)

    result = _run(tool, {"path": "message.txt"})

    assert result.details["ok"] is True
    assert result.details["kind"] == "text"
    payload = json.loads(result.content[0].text)
    assert payload["ok"] is True
    assert payload["content"] == "hello tool"


def test_file_read_tool_error_envelope(tmp_path: Path):
    tool = FileReadTool(workspace_root=str(tmp_path))
    result = _run(tool, {"path": "missing.txt"})

    assert result.details == {
        "ok": False,
        "error": "path_not_found",
        "message": "file not found: missing.txt",
        "path": "missing.txt",
    }
    assert result.content[0].text == "file_read error: file not found: missing.txt"


def test_read_text_line_window(tmp_path: Path):
    (tmp_path / "notes.txt").write_text("line1\nline2\nline3\nline4\n", encoding="utf-8")

    result = read("notes.txt", workspace_root=str(tmp_path), line_start=2, line_count=2)

    assert result["ok"] is True
    assert result["content"] == "line2\nline3"
    assert result["metadata"]["line_count"] == 4
    assert result["metadata"]["requested_line_start"] == 2
    assert result["metadata"]["requested_line_count"] == 2
    assert result["metadata"]["returned_line_start"] == 2
    assert result["metadata"]["returned_line_end"] == 3


def test_read_text_query_returns_bounded_context(tmp_path: Path):
    (tmp_path / "notes.txt").write_text(
        "\n".join(
            [
                "chapter one",
                "nothing here",
                "the phrase endopsychic myths appears here",
                "follow-up detail",
                "separator",
                "another mention of endopsychic myths later",
                "closing line",
            ]
        ),
        encoding="utf-8",
    )

    result = read(
        "notes.txt",
        workspace_root=str(tmp_path),
        query="endopsychic myths",
        context_lines=1,
        max_matches=2,
    )

    assert result["ok"] is True
    assert result["metadata"]["query"] == "endopsychic myths"
    assert result["metadata"]["query_match_count"] == 2
    assert result["metadata"]["query_matched_line_numbers"] == [3, 6]
    assert "--- match 1 lines 2-7 ---" in result["content"]
    assert "3: the phrase endopsychic myths appears here" in result["content"]
    assert "6: another mention of endopsychic myths later" in result["content"]


def test_read_text_query_reports_no_match(tmp_path: Path):
    (tmp_path / "notes.txt").write_text("alpha\nbeta\ngamma\n", encoding="utf-8")

    result = read("notes.txt", workspace_root=str(tmp_path), query="delta")

    assert result["ok"] is True
    assert result["metadata"]["query_match_count"] == 0
    assert result["content"] == "No matches found for query: delta"


def test_read_pdb_includes_atom_hints_and_serializes_metadata_first(tmp_path: Path):
    pdb_text = """HEADER demo
REMARK x
ATOM      1  N   GLY A   1      11.104  13.207   9.199
ATOM      2  CA  GLY A   1      12.560  13.400   9.800
ATOM      3  C   GLY A   1      13.000  12.100  10.300
"""
    (tmp_path / "structure.pdb").write_text(pdb_text, encoding="utf-8")

    result = read("structure.pdb", workspace_root=str(tmp_path), max_chars=32)

    assert result["ok"] is True
    assert result["truncated"] is True
    assert result["metadata"]["first_atom_line"] == 3
    assert result["metadata"]["atom_preview"][:2] == [
        "ATOM      1  N   GLY A   1      11.104  13.207   9.199",
        "ATOM      2  CA  GLY A   1      12.560  13.400   9.800",
    ]

    rendered = json.dumps(result, ensure_ascii=False)
    assert rendered.index('"metadata"') < rendered.index('"content"')
